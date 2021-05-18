from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import string

allDict = {}
workList = []
expenseList = []

feduro = 'FEDURO'
felipe = 'FELIPE MOTTA'
pascual = 'PRODUCTOS ALIMENTICIOS PANAMA'
estrella_azul = 'MPOS INDUSTRIAS LACTEA'
varela = 'VARELA'
kenny = 'KENNY'
julissa = 'JULISSA'

# Create three dictionaries for cleaning
def clean_lists():
  allDict.clear()
  workList.clear()
  expenseList.clear()

# Function to calculate the total with a given list
def calculate_total(myList):
  total = 0   
  index = 1
  while index < len(myList):
    total += myList[index]
    index += 2
  return total  

# Function to clean the workbook, only obtaining the entity and amount paid
def clean_workbook(wrkbk):
  # load workbook and get the active sheet
  wb = load_workbook(wrkbk)
  ws = wb.active
  # Delete the innecessary columns
  if str(ws['B1'].value) != 'cleared':
    ws.delete_cols(1,4)
    ws.delete_cols(3,4)
  ws['B1'] = 'cleared'
  wb.save(wrkbk)

# Function that creates the results workbook
def create_workbook():
  wb = Workbook()
  ws = wb.active
  ws.title = "Totals"
  wb.save('files/results.xlsx')

# Function that creates a new workbook and saves the work total, expense total, and final total
def workbook_results(name):
  work_total = 0
  expense_total = 0
  total = 0
  # Create a new xlsx file
  wb = load_workbook('files/results.xlsx')
  ws = wb.active
  ws.append(name)
  # Obtain the work total and the expense total using the calculate_total function
  work_total = calculate_total(workList)
  expense_total = calculate_total(expenseList)
  # Append the results to the new file
  ws.append(['Pagos Trabajo', work_total])
  ws.append(['Gastos', expense_total])
  # Sum up the expenses and work totals to obtain the overall total
  total = expense_total + work_total
  ws.append(['Total', total])
  # Save the to the xlsx file
  wb.save('files/results.xlsx')

# Function that returns true or false whether the entity is for work or not
def isWork(myString):
  if felipe in str(myString) or pascual in str(myString) or estrella_azul in str(myString) or varela in str(myString) or feduro in str(myString):
    return True
  else:
    return False

# Function that returns the index of the highest number of transactions for looping
def get_trans_qty(wrkbk):
  wb = load_workbook(wrkbk)
  ws = wb.active
  index = 9
  while index < 500:
    if str(ws['B' + str(index)].value) == 'None':  
      break
    else:
      index += 1
  return index

# Function that calculates the sum inside the workbook to obtain the grand total
def sum_of_totals(rslts):
  wb = load_workbook(rslts)
  ws = wb.active
  ws['A9'].value = 'Grand Total'
  ws['B9'].value = '=SUM(B4, B8)'
  wb.save('files/results.xlsx')

# Function that does all of the work for one workbook
def entire_process(wrkbk):
  clean_workbook(wrkbk)
  trans = get_trans_qty(wrkbk)
  wb = load_workbook(wrkbk)
  ws = wb.active
  # Obtain every row (company and amount paid) and store it in a dictionary
  # for row in range (9, 50):
  for row in range (9, trans):
    allDict['company' + str(row)] = ws['B' + str(row)].value
    allDict['paid' + str(row)] = ws['C' + str(row)].value
  # Turn the dictionary into a list
  value_list = list(allDict.values())
  # Create an index to loop through each item in the list, and if it contains any of the key words, then append them to the work dictionary,
  # otherwise, append them to the expense dictionary
  index = 0
  while index < len(value_list):
    if isWork(str(value_list[index])) == True:
      workList.append(value_list[index])
      workList.append(value_list[index+1])
      index += 2
    else: 
      if str(value_list[index+1]) != 'None':
        expenseList.append(value_list[index])
        expenseList.append(value_list[index+1])
      index += 2
  # Save the to the xlsx file
  wb.save(wrkbk)

def main():
  # Subtitute the workbook name to apply to the rest of the code
  myWorkbook = 'files/dad_credit_card.xlsx'
  entire_process(myWorkbook)
  create_workbook()
  workbook_results(['Dad'])
  myWorkbook = 'files/kenny_credit_card.xlsx'
  clean_lists()
  entire_process(myWorkbook)
  workbook_results(['Kenny'])
  sum_of_totals('files/results.xlsx')

if __name__ == "__main__":
    main()