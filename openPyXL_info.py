from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Load a workbook
# wb = load_workbook("kenny.xlsx")

# Create a workbook
wb = Workbook()

# Obtain the active worksheet
ws = wb.active

# Change to a different sheet
# ws = wb['Sheet2']

# Set the worksheet title
ws.title = "testing"

# Append a row of data, separated by commas for each cell
# ws.append(['hello', 'my', 'name', 'is', 'kenny'])

# Merge cells in a range
# ws.merge_cells('A1:D1')

# Unmerge cells in a range
# ws.unmerge_cells('A1:D1')

# Insert rows, specify which in which row you want to insert
# ws.insert_rows(7)

# Delete rows, specify which in which row you want to delete
# ws.delete_rows(7)

# Insert cols, specify which in which col you want to insert
# ws.insert_cols(2)

# Delete cols, specify which in which col you want to delete
# ws.delete_cols(2)

# Move cell range up(-)/down(+) and left(-)/right(+)
# ws.move_range("C1:D11", rows=2, cols=2)

# Iterate over the rows from 1 to 11 exclusive
for row in range (1,11):
  # Iterate over the cols from 1 to 5 exclusive
  for col in range (1, 5):
    # Obtain the letter of the col
    char = get_column_letter(col)
    # Set the value to the letter + row number
    ws[char + str(row)] = char + str(row)

# Save the workbook and give it a name
wb.save('kenny.xlsx')